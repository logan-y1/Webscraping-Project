from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
from urllib.request import urlopen, Request


url = "https://crypto.com/price"
# Request in case 404 Forbidden error
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3"
}

req = Request(url, headers=headers)
webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, "html.parser")


table_rows = soup.findAll("td")


# EXCEL

wb = xl.Workbook()

ws = wb.active

ws.title = "Crypto Currencies"


myfont = Font(name="Chalkboard", size=24, bold=True, italic=False)
myfont2 = Font(name="Chalkboard", size=14, bold=True, italic=False)

maxC = ws.max_column
maxR = ws.max_row


ws["A1"] = "Top Five Crypto Currencies"
ws["A1"].font = myfont
ws["A3"] = "Name"
ws["A3"].font = myfont2
ws["B3"] = "Price"
ws["B3"].font = myfont2
ws["C3"] = "Change"
ws["C3"].font = myfont2
ws.merge_cells("A1:B1")

write_row = 2
write_colA = 1
write_colB = 2
write_colC = 3
write_colD = 4

for currentrow in ws.iter_rows(min_row=2, max_row=maxR, max_col=maxC):
    td = row.findAll("td")
    name = td[1].text
    price = td[2].text
    change = td[3].text

    dollar_change = float(price) * float(change)

    ws.cell(write_row, write_colA).value = name
    ws.cell(write_row, write_colB).value = price
    ws.cell(write_row, write_colC).value = change


wb.save("cryptocurrencies.xlsx")


import keys_copy
from twilio.rest import Client


client = Client(keys_copy.accountSID, keys_copy.authToken)
TwilioNumber = "(507) 710-7908"
mycellphone = "4692363979"

if dollar_change >= 5:
    message = "Your crypto currency has increased by $5.00 in the last 24hr!"
    textmessage = client.messages.create(
        to=mycellphone, from_=TwilioNumber, body=message
    )
